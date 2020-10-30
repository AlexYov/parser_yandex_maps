import requests
import json
from tkinter import *
from openpyxl import Workbook
import PySimpleGUI as sg

sg.theme('Default1')

wb = Workbook()
ws = wb.active
ws = wb.create_sheet("база",0)

prog = Tk()
prog.geometry("400x150")

name_org = Entry(prog)
name_city = Entry(prog)
results = Entry(prog)

label_name_org = Label(text='Примерное название организации')
label_name_city = Label(text='Город/Область/Регион')
label_results = Label(text='Сколько результатов вернуть? (Max 500)')

name_org.grid(row=0, column=1)
label_name_org.grid(row=0, column=0)

name_city.grid(row=1, column=1)
label_name_city.grid(row=1, column=0)

results.grid(row=2, column=1)
label_results.grid(row=2, column=0)

name_org.focus_set()
name_city.focus_set()
results.focus_set()

def func1():
    response = requests.get('https://search-maps.yandex.ru/v1/?text={}, {}&type=biz&lang=ru_RU&results={}&apikey=здесь API Яндекс аккаунта'.format(name_org.get(),name_city.get(),results.get()))

    with open("organizations.json", "w") as write_file:
        result = json.dumps(response.json(), ensure_ascii=True)
        write_file.write(result)

    with open('organizations.json') as f:
        file_content = f.read()
        finish = json.loads(file_content)

        names = []
        descriptions = []
        urls = []
        phones = []
        сategories = []
        for i in range(len(finish['features'])):
            try:
                names.append(finish['features'][i]['properties']['name'])
            except:
                names.append('-')

            try:
                descriptions.append(finish['features'][i]['properties']['description'])
            except:
                descriptions.append('-')

            try:
                urls.append(finish['features'][i]['properties']['CompanyMetaData']['url'])
            except:
                urls.append('-')

            try:
                CompanyMetaData_Phones=[]
                for y in range(len(finish['features'][i]['properties']['CompanyMetaData']['Phones'])):
                    CompanyMetaData_Phones.append(finish['features'][i]['properties']['CompanyMetaData']['Phones'][y]['formatted'])
                phones.append(str(CompanyMetaData_Phones))
            except:
                phones.append(' ')


            try:
                names_Categories = []
                for o in range(len(finish['features'][i]['properties']['CompanyMetaData']['Categories'])):
                    names_Categories.append(finish['features'][i]['properties']['CompanyMetaData']['Categories'][o]['name'])
                сategories.append(str(names_Categories))
            except:
                сategories.append('-')

        row1 = 1
        row2 = 1
        row3 = 1
        row4 = 1
        row5 = 1
        row6 = 1

        for name in names:
            row1+=1
            ws['A'+str(row1)] = name

        for description in descriptions:
            row2+=1
            ws['B'+str(row2)] = description

        for url in urls:
            row3+=1
            ws['C'+str(row3)] = url

        for phone in phones:
            row4+=1
            ws['D'+str(row4)] = phone

        for сategory in сategories:
            row5+=1
            ws['E'+str(row5)] = сategory

        layout = [  [sg.Text('Имя файла БД'), sg.InputText()],
        [sg.Text("Путь сохранения БД"), sg.InputText(),sg.FolderBrowse('Обзор')],
        [sg.Text('Имя файла с телефонами'), sg.InputText()],[sg.Text("Путь сохранения телефонов"), sg.InputText(),sg.FolderBrowse('Обзор')],
                    [sg.Save('Сохранить'),sg.Button('Отмена')] ]

        window = sg.Window('Проводник', layout)

        while True:

            event, values = window.read()

            if event == 'Сохранить':

                if values['Обзор'] == '':
                    event, values=sg.Window('Предупреждение',[[sg.Text('Выберите путь сохранения БД')],[sg.Button('Закрыть')]]).read(close=True)

                elif values[0] == '':
                    event, values=sg.Window('Предупреждение',[[sg.Text('Введите имя БД')],[sg.Button('Закрыть')]]).read(close=True)

                elif values[2] == '':
                    event, values=sg.Window('Предупреждение',[[sg.Text('Введите имя файла с телефонами')],[sg.Button('Закрыть')]]).read(close=True)

                elif values[3] == '':
                    event, values=sg.Window('Предупреждение',[[sg.Text('Выберите путь сохранения файла с телефонами')],[sg.Button('Закрыть')]]).read(close=True)

                else:
                    wb.save(values['Обзор']+'/'+values[0]+'.xlsx')
                    with open(values[3]+'/'+values[2]+'.txt','w') as filetxt:
                        for new_phones in phones:
                            filetxt.write(new_phones.replace('[',' ').replace(']',' ').replace(',','\n').replace("'",'').replace("'",'')+'\n')
                    break

            if event == sg.WIN_CLOSED or event == 'Отмена': 
                break

        window.close()

        label_finish = Label(text='Готово. Проверь файл.')
        label_finish.grid(row=4,column=0)

btn = Button(text='Выполнить поиск',command = func1)
btn.grid(row=3,column=1)
prog.mainloop()
